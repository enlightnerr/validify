from fastapi import FastAPI, File, UploadFile
from fastapi.middleware.cors import CORSMiddleware
import pdfplumber
import openpyxl
from PIL import Image, ImageDraw
import io
import base64

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.post("/extract")
async def extract_pdf_data(file: UploadFile = File(...)):
    # Read file into memory
    content = await file.read()
    
    tables_data = []
    text_data = []
    images_base64 = []
    
    # Extract text and tables using pdfplumber
    with pdfplumber.open(io.BytesIO(content)) as pdf:
        for i, page in enumerate(pdf.pages):
            text = page.extract_text()
            if text:
                text_data.append({"page": i + 1, "text": text})
            
            tables = page.extract_tables()
            if tables:
                tables_data.append({"page": i + 1, "tables": tables})
                
            # Convert page to image using pdfplumber's pypdfium2 integration
            # which produces a PIL image.
            img = page.to_image(resolution=144).original
            img_byte_arr = io.BytesIO()
            img.save(img_byte_arr, format='PNG')
            b64_img = base64.b64encode(img_byte_arr.getvalue()).decode("utf-8")
            images_base64.append(f"data:image/png;base64,{b64_img}")

    return {
        "text": text_data,
        "tables": tables_data,
        "images": images_base64,
        "pageCount": len(images_base64)
    }

@app.post("/extract-excel")
async def extract_excel_data(file: UploadFile = File(...)):
    content = await file.read()
    
    tables_data = []
    text_data = []
    images_base64 = []
    
    try:
        # data_only=True evaluates the formulas into their end state natively
        xls = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        for sheet_name in xls.sheetnames:
            sheet = xls[sheet_name]
            
            data = []
            for row in sheet.iter_rows(values_only=True):
                clean_row = [str(cell) if cell is not None else "" for cell in row]
                data.append(clean_row)
            
            columns = data[0] if len(data) > 0 else []
            rows = data[1:] if len(data) > 1 else []
            
            text_str = ""
            for r in data[:50]: # limit text scope for MVP basic logic
                text_str += " | ".join(r) + "\n"
                
            text_data.append({"page": sheet_name, "text": text_str})
            tables_data.append({"page": sheet_name, "tables": [{"data": rows, "columns": columns}]})
            
            # Use native pure-python PIL to avoid C-level DLL hooks
            img = Image.new('RGBA', (1000, 800), color=(255, 255, 255, 255))
            draw = ImageDraw.Draw(img)
            
            # Simple placeholder map layout for visual pixel processing
            draw.text((20, 20), text_str[:2500], fill=(0, 0, 0, 255))
            
            img_byte_arr = io.BytesIO()
            img.save(img_byte_arr, format='PNG')
            b64_img = base64.b64encode(img_byte_arr.getvalue()).decode("utf-8")
            images_base64.append(f"data:image/png;base64,{b64_img}")
            
    except Exception as e:
        print(f"Error parsing excel: {str(e)}")
        
    return {
        "text": text_data,
        "tables": tables_data,
        "images": images_base64,
        "pageCount": len(images_base64)
    }

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
